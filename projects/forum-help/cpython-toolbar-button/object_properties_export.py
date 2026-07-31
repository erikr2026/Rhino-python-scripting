"""
Object properties -> Excel export (Python 3 / CPython, openpyxl).

ENGINE: This script targets Rhino 8's CPython 3 engine. Run it via the
ScriptEditor command, or from a toolbar button/alias using:

    -_ScriptEditor _Run "C:\\Path\\object_properties_export.py"

Do NOT wire it to a `RunPythonScript`/`!_-RunPythonScript` button. That
legacy command always launches IronPython (Python 2), which cannot import
openpyxl (a CPython-only package here) -- this is the exact failure Luca
hit on the McNeel forum ("Better object properties export", 2026-06-17,
https://discourse.mcneel.com/t/better-object-properties-export/220056),
and the `-_ScriptEditor _Run` fix was confirmed by McNeel staff in the
companion thread
https://discourse.mcneel.com/t/running-a-cpython-script-from-a-rhino-button-icon-instead-of-ironpython/220021.
See README.md in this folder for the full toolbar-button setup and the
gotchas (absolute paths required, briefly opens the ScriptEditor UI, no
confirmed argument-passing syntax).

WHAT THIS DOES
Walks every point/curve/surface/polysurface object in the active document
and writes one row per object to an .xlsx workbook:
    Name | Layer | Object Type | Area (m^2) | Centroid X (m) | Centroid Y (m) | Centroid Z (m)

- Units: areas and centroid coordinates are always converted to meters
  regardless of the document's current model unit, using
  rs.UnitScale(target, source) -- this directly answers Luca's "always
  export in meters" ask.
- Centroid coordinates are written as three separate numeric columns
  (X/Y/Z), not a combined string -- this directly answers Luca's "split
  centroid coordinates" ask.
- Cells are written as native Python floats/ints (not strings), so Excel
  recognizes them as numbers for sorting/formatting/formulas -- this
  answers Luca's "values recognized as numbers" ask. A numeric format
  string is also applied to the numeric columns.

UNVERIFIED / OUT OF SCOPE
- Solid volume export (e.g. for closed polysurfaces) is NOT included.
  rhinoscriptsyntax has no SurfaceVolume/SurfaceVolumeCentroid function
  (confirmed by reading the actual rhinoscriptsyntax source on GitHub,
  2026-07-31 -- only SurfaceArea/SurfaceAreaCentroid/SurfaceAreaMoments
  exist there). Getting volume would require dropping to raw RhinoCommon
  (Rhino.Geometry.VolumeMassProperties.Compute on a solid Brep) -- add
  that separately if needed, it's not implemented here.
- Layer/sublayer split into separate columns (another item from Luca's
  original ask) is not implemented: rs.ObjectLayer returns the full
  layer path (e.g. "Hull::Plating") as one string. Splitting on "::"
  into separate columns is straightforward to add but was left out here
  to keep this example focused -- see the "layer_path.split('::')" note
  in the code if you need it.
- Working directory when launched via `-_ScriptEditor _Run` is not
  confirmed by either forum thread, so the output path is built from
  this script's own folder (via __file__), not a relative path.
"""

import os
import datetime

import rhinoscriptsyntax as rs
import openpyxl
from openpyxl.styles import Font

# Object types worth reporting on (bit-coded, per rhinoscriptsyntax.ObjectType):
# 4 = Curve, 8 = Surface, 16 = Polysurface
GEOMETRY_TYPE_FILTER = 4 | 8 | 16

# Target export unit: 4 = Meters (rs.UnitSystem numeric code).
METERS = 4


def get_export_scale():
    """Scale factor to convert the document's current model units to meters."""
    return rs.UnitScale(METERS)


def area_and_centroid_m(object_id, scale):
    """
    Return (area_m2, x_m, y_m, z_m) for a curve or surface/polysurface,
    or None if the object has no meaningful area (e.g. open curve).
    Coordinates/areas are converted to meters using `scale`.
    """
    obj_type = rs.ObjectType(object_id)

    if obj_type == 4:  # Curve
        if not rs.IsCurveClosed(object_id) or not rs.IsCurvePlanar(object_id):
            return None
        area_info = rs.CurveAreaCentroid(object_id)
    else:  # Surface (8) or Polysurface (16)
        area_info = rs.SurfaceAreaCentroid(object_id)

    if not area_info:
        return None

    area, _error_bound = rs.CurveArea(object_id) if obj_type == 4 else rs.SurfaceArea(object_id)
    centroid = area_info[0]

    area_m2 = area * (scale ** 2)
    x_m = centroid.X * scale
    y_m = centroid.Y * scale
    z_m = centroid.Z * scale
    return area_m2, x_m, y_m, z_m


def object_type_label(obj_type):
    return {4: "Curve", 8: "Surface", 16: "Polysurface"}.get(obj_type, "Other")


def main():
    object_ids = rs.ObjectsByType(GEOMETRY_TYPE_FILTER, select=False, state=0)
    if not object_ids:
        print("No curve/surface/polysurface objects found in the document.")
        return

    scale = get_export_scale()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Object Properties"

    headers = [
        "Name", "Layer", "Object Type",
        "Area (m^2)", "Centroid X (m)", "Centroid Y (m)", "Centroid Z (m)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    numeric_format = "0.000"
    rows_written = 0

    for object_id in object_ids:
        name = rs.ObjectName(object_id) or ""
        layer_path = rs.ObjectLayer(object_id) or ""
        # To split layer/sublayer into separate columns instead, use:
        #   layer_parts = layer_path.split("::")
        obj_type = rs.ObjectType(object_id)

        result = area_and_centroid_m(object_id, scale)
        if result is None:
            # No area (e.g. open curve) -- still log identity columns, blank numerics.
            ws.append([name, layer_path, object_type_label(obj_type), None, None, None, None])
            continue

        area_m2, x_m, y_m, z_m = result
        row = [name, layer_path, object_type_label(obj_type), area_m2, x_m, y_m, z_m]
        ws.append(row)

        row_idx = ws.max_row
        for col_idx in range(4, 8):  # Area, X, Y, Z columns
            ws.cell(row=row_idx, column=col_idx).number_format = numeric_format

        rows_written += 1

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(14, len(header) + 2)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    out_path = os.path.join(script_dir, "object_properties_{}.xlsx".format(timestamp))

    wb.save(out_path)
    print("Exported {} object(s) with area data ({} total scanned) to:\n{}".format(
        rows_written, len(object_ids), out_path))


if __name__ == "__main__":
    main()
